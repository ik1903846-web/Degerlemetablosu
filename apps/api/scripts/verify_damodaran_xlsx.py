#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Damodaran XLSX Verification — Faz B1 Adim 2B
============================================

parameters.json icindeki KNOWN_PARAMS degerlerini fetch'lenen
xlsx icerikleriyle karsilastir. Windows cp1252 uyumlu.
"""

import io
import json
import sys
from pathlib import Path

# Windows console UTF-8 fix
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer, encoding="utf-8", errors="replace"
    )
    sys.stderr = io.TextIOWrapper(
        sys.stderr.buffer, encoding="utf-8", errors="replace"
    )

try:
    from openpyxl import load_workbook
except ImportError:
    print("HATA: openpyxl gerekli. Kurmak: pip install openpyxl")
    sys.exit(2)


DATA_DIR = (
    Path(__file__).resolve().parent.parent
    / "data" / "damodaran" / "2026_05_09"
)


def load_parameters_json() -> dict:
    path = DATA_DIR / "parameters.json"
    if not path.exists():
        print(f"HATA: {path} bulunamadi. Adim 1 fetch calistirilmadi.")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def find_col(headers: dict, *keywords) -> int | None:
    """Header dict'inde tum keywords iceren kolonun index'ini bul."""
    for h, idx in headers.items():
        if all(kw in h for kw in keywords):
            return idx
    return None


def build_header_map(ws, header_row: int) -> dict:
    """Header satirini normalize edilmis dict'e cevir."""
    headers = {}
    for col_idx, cell in enumerate(ws[header_row]):
        if cell.value:
            normalized = str(cell.value).strip().lower()
            headers[normalized] = col_idx
    return headers


def find_country_row(ws, header_row: int, country: str) -> tuple | None:
    """Verilen ulke icin row'u bul."""
    for row in ws.iter_rows(
        min_row=header_row + 1, max_row=300, values_only=True
    ):
        if row[0] and country.lower() in str(row[0]).lower():
            return row
    return None


def parse_ctryprem_turkey() -> dict:
    """
    Turkey satirini header-based dynamic mapping ile parse et.
    Tax rate ayri sheet'ten ('Country Tax Rates') okunur.
    """
    path = DATA_DIR / "ctryprem.xlsx"
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {len(sheet_names)} sheet")

    # 1) "ERPs by country" sheet'i bul
    target_sheet = None
    for sn in sheet_names:
        sn_lower = sn.lower()
        if "erp" in sn_lower and "country" in sn_lower:
            target_sheet = sn
            break
    if target_sheet is None:
        raise ValueError(
            f"'ERPs by country' sheet bulunamadi. "
            f"Mevcut: {sheet_names}"
        )
    print(f"  Target sheet: {target_sheet}")
    ws = wb[target_sheet]

    # 2) Header satirini bul (2-15 arasi tara, row 1 title olabilir)
    # Sıkı kriter: "country" + ("rating" OR "moody") zorunlu.
    # "default"/"premium" eklenirse title satirina false positive olur.
    header_row = None
    for row_idx in range(2, 16):
        row_values = [(cell.value or "") for cell in ws[row_idx]]
        row_str = " ".join(str(v) for v in row_values).lower()
        # Bos satirlari atla
        non_empty = sum(1 for v in row_values if v)
        if non_empty < 3:
            continue
        # Siki kriter: country + (rating veya moody)
        if "country" in row_str and (
            "rating" in row_str or "moody" in row_str
        ):
            header_row = row_idx
            print(f"  Header row: {row_idx}")
            print(f"  Headers: {row_values[:10]}")
            break

    if header_row is None:
        raise ValueError(
            "Header satiri bulunamadi (2-15 satir tarandi). "
            "Kriter: 'country' + ('rating' veya 'moody'), "
            "min 3 dolu hucre."
        )

    # 3) Header'lari dynamic dict'e cevir
    headers = build_header_map(ws, header_row)

    # 4) Kolon indekslerini header isimleriyle bul
    col_rating = find_col(headers, "moody", "rating") \
        or find_col(headers, "rating")
    col_default = find_col(headers, "default", "spread") \
        or find_col(headers, "rating-based")
    col_total_erp = find_col(headers, "total", "equity", "risk", "premium")
    col_crp = find_col(headers, "country", "risk", "premium")
    col_cds = find_col(headers, "sovereign", "cds")
    # ERP based on CDS: ikinci "Total Equity Risk Premium" veya
    # "ERP based on" iceren kolon
    col_erp_cds = find_col(headers, "erp", "based") \
        or find_col(headers, "premium2") \
        or find_col(headers, "premium 2")

    print(f"  Column mapping:")
    print(f"    rating:        col={col_rating}")
    print(f"    default:       col={col_default}")
    print(f"    total_erp:     col={col_total_erp}")
    print(f"    crp:           col={col_crp}")
    print(f"    cds:           col={col_cds}")
    print(f"    erp_cds:       col={col_erp_cds}")

    if any(c is None for c in [
        col_rating, col_default, col_total_erp, col_crp
    ]):
        raise ValueError(
            f"Kritik kolon bulunamadi. Headers: {list(headers.keys())}"
        )

    # 5) Turkey row'u oku
    turkey_row = find_country_row(ws, header_row, "Turkey")
    if turkey_row is None:
        raise ValueError("Turkey satiri bulunamadi")
    print(f"  Turkey row sample: {turkey_row[:8]}")

    # 6) Tax rate ayri sheet'ten
    tax_rate = None
    tax_sheet_name = None
    for sn in sheet_names:
        if "tax" in sn.lower() and "rate" in sn.lower():
            tax_sheet_name = sn
            break
    if tax_sheet_name:
        print(f"  Tax sheet: {tax_sheet_name}")
        tax_ws = wb[tax_sheet_name]
        # Tax sheet header bul (genelde Country + Tax Rate kolonlari)
        tax_header_row = None
        for row_idx in range(1, 10):
            row_values = [
                (cell.value or "") for cell in tax_ws[row_idx]
            ]
            row_str = " ".join(str(v) for v in row_values).lower()
            if "country" in row_str and "tax" in row_str:
                tax_header_row = row_idx
                break
        if tax_header_row:
            tax_headers = build_header_map(tax_ws, tax_header_row)
            tax_col = find_col(tax_headers, "tax", "rate") \
                or find_col(tax_headers, "corporate", "tax")
            if tax_col is not None:
                turkey_tax_row = find_country_row(
                    tax_ws, tax_header_row, "Turkey"
                )
                if turkey_tax_row:
                    tax_rate = turkey_tax_row[tax_col]
                    print(f"  Tax rate (Turkey): {tax_rate}")

    return {
        "country_label": turkey_row[0],
        "rating": turkey_row[col_rating],
        "default_spread": turkey_row[col_default],
        "crp": turkey_row[col_crp],
        "total_erp": turkey_row[col_total_erp],
        "tax_rate": tax_rate,
        "cds": turkey_row[col_cds] if col_cds is not None else None,
        "erp_cds": (
            turkey_row[col_erp_cds] if col_erp_cds is not None else None
        ),
    }


def compare(known: dict, parsed: dict) -> list[str]:
    discrepancies = []
    TOLERANCE = 0.0005

    checks = [
        ("rating",
         known["turkey"]["rating"],
         parsed["rating"], "exact"),
        ("default_spread",
         known["turkey"]["default_spread"],
         parsed["default_spread"], "numeric"),
        ("crp",
         known["turkey"]["crp"],
         parsed["crp"], "numeric"),
        ("total_erp",
         known["turkey"]["total_erp_lambda1"],
         parsed["total_erp"], "numeric"),
        ("tax_rate",
         known["turkey"]["tax_rate"],
         parsed["tax_rate"], "numeric"),
        ("cds",
         known["turkey"]["sovereign_cds"],
         parsed["cds"], "numeric"),
    ]

    for name, k_val, p_val, mode in checks:
        if p_val is None:
            discrepancies.append(
                f"  WARN {name}: parsed=None (xlsx'te yok)"
            )
            continue
        if mode == "exact":
            if str(k_val).strip() != str(p_val).strip():
                discrepancies.append(
                    f"  FAIL {name}: known={k_val!r}, "
                    f"parsed={p_val!r}"
                )
            else:
                print(f"  PASS {name}: {k_val} = {p_val}")
        elif mode == "numeric":
            try:
                k_num = float(k_val)
                p_num = float(p_val)
                if abs(k_num - p_num) > TOLERANCE:
                    discrepancies.append(
                        f"  FAIL {name}: known={k_num:.4f}, "
                        f"parsed={p_num:.4f}, "
                        f"delta={abs(k_num-p_num):.4f}"
                    )
                else:
                    print(
                        f"  PASS {name}: known={k_num:.4f}, "
                        f"parsed={p_num:.4f} (tolerans icinde)"
                    )
            except (TypeError, ValueError):
                discrepancies.append(
                    f"  WARN {name}: numeric parse fail "
                    f"known={k_val!r} parsed={p_val!r}"
                )

    return discrepancies


def main() -> int:
    print("=" * 60)
    print("Damodaran XLSX Verification - Faz B1 Adim 2B")
    print("=" * 60)

    print("\n[1] parameters.json yukleniyor...")
    known = load_parameters_json()
    print(f"    fetch_date: {known.get('fetch_date')}")

    print("\n[2] ctryprem.xlsx -> Turkey satiri parse...")
    try:
        parsed = parse_ctryprem_turkey()
    except Exception as e:
        print(f"    HATA: {e}")
        return 1

    print("\n[3] KNOWN_PARAMS vs xlsx karsilastirma...")
    discrepancies = compare(known, parsed)

    print("\n" + "=" * 60)
    if discrepancies:
        print(f"FAIL: {len(discrepancies)} SAPMA TESPIT EDILDI:")
        for d in discrepancies:
            print(d)
        print("\nAUDIT DERINLESTIRME GEREKLI.")
        return 1
    else:
        print("PASS: TUM PARAMETRELER UYUMLU (xlsx vs KNOWN_PARAMS)")
        print("  Adim 2C'ye gecilebilir (production diff).")
        return 0


if __name__ == "__main__":
    sys.exit(main())
